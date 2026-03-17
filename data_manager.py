
import json
import os
from datetime import datetime
from database import get_db_connection, init_database

DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')

class DataManager:
    def __init__(self):
        # Inicializar banco de dados
        try:
            init_database()
        except Exception as e:
            print(f"Erro ao inicializar banco: {e}")
            raise

    def get_candidates(self):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT * FROM candidates ORDER BY id')
        candidates = [dict(row) for row in cur.fetchall()]
        cur.close()
        conn.close()
        return candidates

    def save_candidates(self, candidates):
        # Não usado mais, mas mantido para compatibilidade
        pass

    def add_candidate(self, candidate):
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute('''
            INSERT INTO candidates (nome, justificativa, gestor, periodo, categoria, photo)
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING id, nome, justificativa, gestor, periodo, categoria, photo
        ''', (
            candidate.get('nome'),
            candidate.get('justificativa', ''),
            candidate.get('gestor', ''),
            candidate.get('periodo', 'Q3/2025'),
            candidate.get('categoria', 'Eu Faço a Diferença'),
            candidate.get('photo', 'default_avatar.png')
        ))
        
        new_candidate = dict(cur.fetchone())
        
        # Inicializar contagem de votos
        cur.execute('''
            INSERT INTO votes (candidate_id, count) VALUES (%s, 0)
            ON CONFLICT (candidate_id) DO NOTHING
        ''', (new_candidate['id'],))
        
        conn.commit()
        cur.close()
        conn.close()
        return new_candidate
    
    def import_candidates_preserving_existing(self, new_candidates):
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Obter nomes existentes
        cur.execute('SELECT LOWER(TRIM(nome)) as nome FROM candidates')
        existing_names = {row['nome'] for row in cur.fetchall()}
        
        added_count = 0
        for new_candidate in new_candidates:
            if new_candidate['nome'].lower().strip() not in existing_names:
                # Garantir que photo seja incluído (preservar base64 do backup)
                if 'photo' not in new_candidate or not new_candidate['photo']:
                    new_candidate['photo'] = 'default_avatar.png'
                elif not new_candidate['photo'].startswith('data:image') and new_candidate['photo'] != 'default_avatar.png':
                    # Se não é base64 nem default, usar default
                    new_candidate['photo'] = 'default_avatar.png'
                self.add_candidate(new_candidate)
                added_count += 1
        
        cur.close()
        conn.close()
        return added_count

    def update_candidate(self, candidate_id, updated_data):
        conn = get_db_connection()
        cur = conn.cursor()
        
        set_clause = ', '.join([f"{k} = %s" for k in updated_data.keys()])
        values = list(updated_data.values()) + [candidate_id]
        
        cur.execute(f'UPDATE candidates SET {set_clause} WHERE id = %s', values)
        conn.commit()
        cur.close()
        conn.close()

    def delete_candidate(self, candidate_id):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('DELETE FROM candidates WHERE id = %s', (candidate_id,))
        conn.commit()
        cur.close()
        conn.close()

    def get_votes(self):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT candidate_id, count FROM votes')
        votes = {str(row['candidate_id']): row['count'] for row in cur.fetchall()}
        cur.close()
        conn.close()
        return votes

    def add_vote(self, candidate_id):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('''
            INSERT INTO votes (candidate_id, count) VALUES (%s, 1)
            ON CONFLICT (candidate_id) DO UPDATE SET count = votes.count + 1
        ''', (candidate_id,))
        conn.commit()
        cur.close()
        conn.close()
        self._save_local_backup_votes()

    def remove_vote(self, candidate_id):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('''
            UPDATE votes SET count = GREATEST(0, count - 1)
            WHERE candidate_id = %s
        ''', (candidate_id,))
        conn.commit()
        cur.close()
        conn.close()

    def get_vote_count(self, candidate_id):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT count FROM votes WHERE candidate_id = %s', (candidate_id,))
        result = cur.fetchone()
        cur.close()
        conn.close()
        return result['count'] if result else 0

    def get_voters(self):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT * FROM voters')
        voters = [dict(row) for row in cur.fetchall()]
        cur.close()
        conn.close()
        return voters

    def save_voters(self, voters):
        # Não usado mais
        pass

    def add_voter(self, email, candidate_id):
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Obter categoria do candidato
        cur.execute('SELECT categoria FROM candidates WHERE id = %s', (candidate_id,))
        result = cur.fetchone()
        if not result:
            cur.close()
            conn.close()
            return False
        
        is_leader = 'Líder' in result['categoria']
        
        # Verificar se eleitor já existe
        cur.execute('SELECT * FROM voters WHERE email = %s', (email,))
        existing_voter = cur.fetchone()
        
        if existing_voter:
            if is_leader:
                old_leader_id = existing_voter.get('leader_id')
                if old_leader_id:
                    self.remove_vote(old_leader_id)
                cur.execute('''
                    UPDATE voters SET leader_id = %s, voted_leader = TRUE
                    WHERE email = %s
                ''', (candidate_id, email))
            else:
                old_professional_id = existing_voter.get('professional_id')
                if old_professional_id:
                    self.remove_vote(old_professional_id)
                cur.execute('''
                    UPDATE voters SET professional_id = %s, voted_professional = TRUE
                    WHERE email = %s
                ''', (candidate_id, email))
        else:
            if is_leader:
                cur.execute('''
                    INSERT INTO voters (email, leader_id, voted_leader)
                    VALUES (%s, %s, TRUE)
                ''', (email, candidate_id))
            else:
                cur.execute('''
                    INSERT INTO voters (email, professional_id, voted_professional)
                    VALUES (%s, %s, TRUE)
                ''', (email, candidate_id))
        
        conn.commit()
        cur.close()
        conn.close()
        self._save_local_backup_voters()
        return True

    def get_voter_status(self, email):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT * FROM voters WHERE email = %s', (email,))
        voter = cur.fetchone()
        cur.close()
        conn.close()
        return dict(voter) if voter else None

    def has_voted(self, email, candidate_id):
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute('SELECT categoria FROM candidates WHERE id = %s', (candidate_id,))
        result = cur.fetchone()
        if not result:
            cur.close()
            conn.close()
            return None
        
        is_leader = 'Líder' in result['categoria']
        
        cur.execute('SELECT * FROM voters WHERE email = %s', (email,))
        voter = cur.fetchone()
        
        cur.close()
        conn.close()
        
        if voter:
            return voter.get('leader_id') if is_leader else voter.get('professional_id')
        return None

    def get_config(self):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT * FROM config WHERE id = 1')
        config = dict(cur.fetchone())
        
        # Converter timestamp para string ISO
        if config.get('voting_end_date'):
            config['voting_end_date'] = config['voting_end_date'].isoformat()
        
        cur.close()
        conn.close()
        return config

    def save_config(self, config):
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute('''
            UPDATE config SET 
                voting_end_date = %s,
                voting_active = %s,
                period = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = 1
        ''', (config.get('voting_end_date'), config.get('voting_active'), config.get('period', 'Destaques')))
        
        conn.commit()
        cur.close()
        conn.close()
        self._save_local_backup_config(config)

    def _save_local_backup_config(self, config):
        try:
            os.makedirs(DATA_DIR, exist_ok=True)
            path = os.path.join(DATA_DIR, 'config.json')
            safe_config = {k: v for k, v in config.items() if k != 'id'}
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(safe_config, f, ensure_ascii=False, default=str)
        except Exception as e:
            print(f"Aviso: não foi possível salvar backup local de config: {e}")

    def _save_local_backup_votes(self):
        try:
            os.makedirs(DATA_DIR, exist_ok=True)
            votes = self.get_votes()
            path = os.path.join(DATA_DIR, 'votes.json')
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(votes, f, ensure_ascii=False)
        except Exception as e:
            print(f"Aviso: não foi possível salvar backup local de votes: {e}")

    def _save_local_backup_voters(self):
        try:
            os.makedirs(DATA_DIR, exist_ok=True)
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute('SELECT email, professional_id, leader_id, voted_professional, voted_leader FROM voters')
            voters = [dict(row) for row in cur.fetchall()]
            cur.close()
            conn.close()
            path = os.path.join(DATA_DIR, 'voters.json')
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(voters, f, ensure_ascii=False, default=str)
        except Exception as e:
            print(f"Aviso: não foi possível salvar backup local de voters: {e}")

    def is_voting_active(self):
        config = self.get_config()
        
        if not config.get('voting_active', False):
            return False

        end_date_str = config.get('voting_end_date')
        if not end_date_str:
            return False

        try:
            end_date = datetime.fromisoformat(end_date_str.replace('Z', '+00:00'))
            now = datetime.now()
            
            is_active = now < end_date
            
            if not is_active and config.get('voting_active'):
                config['voting_active'] = False
                self.save_config(config)
            
            return is_active
        except Exception as e:
            print(f"Erro ao verificar data de votação: {e}")
            return False

    def get_results(self):
        candidates = self.get_candidates()
        votes = self.get_votes()

        for candidate in candidates:
            candidate['vote_count'] = votes.get(str(candidate['id']), 0)

        professional = [c for c in candidates if 'Líder' not in c.get('categoria', '')]
        leader = [c for c in candidates if 'Líder' in c.get('categoria', '')]

        professional.sort(key=lambda x: x['vote_count'], reverse=True)
        leader.sort(key=lambda x: x['vote_count'], reverse=True)

        professional_alpha = sorted(professional, key=lambda x: x['nome'])
        leader_alpha = sorted(leader, key=lambda x: x['nome'])

        if len(professional) >= 4 and len(leader) >= 1:
            top_professional = professional[:4]
            top_leader = leader[:1]

            remaining_professional = professional[4:] if len(professional) > 4 else []
            remaining_leader = leader[1:] if len(leader) > 1 else []

            remaining_all = remaining_professional + remaining_leader
            remaining_all.sort(key=lambda x: x['vote_count'], reverse=True)

            if remaining_all:
                fifth = remaining_all[0]
                if 'Líder' in fifth.get('categoria', ''):
                    top_leader.append(fifth)
                else:
                    top_professional.append(fifth)

            top_5 = top_professional + top_leader
        elif len(professional) >= 3 and len(leader) >= 2:
            top_professional = professional[:3]
            top_leader = leader[:2]
            top_5 = top_professional + top_leader
        else:
            all_candidates = professional + leader
            all_candidates.sort(key=lambda x: x['vote_count'], reverse=True)
            top_5 = all_candidates[:5]

        return {
            'top_5': top_5,
            'all_candidates': candidates,
            'professional': professional,
            'leader': leader,
            'professional_alpha': professional_alpha,
            'leader_alpha': leader_alpha
        }

    def reset_voting(self):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('DELETE FROM voters')
        cur.execute('UPDATE votes SET count = 0')
        conn.commit()
        cur.close()
        conn.close()
        self._save_local_backup_votes()
        self._save_local_backup_voters()

    def delete_all_candidates(self):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('DELETE FROM candidates')
        cur.execute('DELETE FROM votes')
        cur.execute('DELETE FROM voters')
        conn.commit()
        cur.close()
        conn.close()
    
    def save_voting_to_history(self, description=''):
        conn = get_db_connection()
        cur = conn.cursor()
        
        candidates = self.get_candidates()
        votes = self.get_votes()
        config = self.get_config()
        
        for candidate in candidates:
            candidate['vote_count'] = votes.get(str(candidate['id']), 0)
        
        cur.execute('''
            INSERT INTO voting_history (description, config, candidates, total_voters, total_votes)
            VALUES (%s, %s, %s, %s, %s)
        ''', (
            description or f"Votação encerrada em {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            json.dumps(config),
            json.dumps(candidates),
            len(self.get_voters()),
            sum(votes.values())
        ))
        
        conn.commit()
        cur.close()
        conn.close()
    
    def get_voting_history(self):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT * FROM voting_history ORDER BY timestamp DESC')
        history = [dict(row) for row in cur.fetchall()]
        
        # Converter timestamps para strings
        for entry in history:
            if entry.get('timestamp'):
                entry['timestamp'] = entry['timestamp'].isoformat()
        
        cur.close()
        conn.close()
        return history

    def delete_history_entry(self, entry_id):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('DELETE FROM voting_history WHERE id = %s', (entry_id,))
        conn.commit()
        cur.close()
        conn.close()

    def import_history_from_excel(self, file_bytes):
        import openpyxl
        import io
        from collections import defaultdict

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows())]

        col_map = {}
        for idx, h in enumerate(headers):
            h_lower = h.lower()
            if 'colaborador' in h_lower:
                col_map['nome'] = idx
            elif 'justificativa' in h_lower:
                col_map['justificativa'] = idx
            elif 'gestor' in h_lower:
                col_map['gestor'] = idx
            elif 'periodo' in h_lower or 'período' in h_lower:
                col_map['periodo'] = idx
            elif 'categoria' in h_lower:
                col_map['categoria'] = idx
            elif 'área' in h_lower or 'area' in h_lower:
                col_map['area'] = idx
            elif 'coment' in h_lower:
                col_map['comentario'] = idx

        by_period = defaultdict(list)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            def get(key):
                idx = col_map.get(key)
                return str(row[idx]).strip() if idx is not None and row[idx] else ''
            nome = get('nome')
            if not nome:
                continue
            periodo = get('periodo') or 'Sem período'
            by_period[periodo].append({
                'nome': nome,
                'justificativa': get('justificativa'),
                'gestor': get('gestor'),
                'categoria': get('categoria'),
                'area': get('area'),
                'comentario': get('comentario'),
                'vote_count': 0,
                'photo': 'default_avatar.png'
            })

        conn = get_db_connection()
        cur = conn.cursor()
        inserted = 0
        for periodo, candidates in by_period.items():
            config_data = {'period': periodo, 'source': 'excel_import'}
            cur.execute('''
                INSERT INTO voting_history (description, config, candidates, total_voters, total_votes)
                VALUES (%s, %s, %s, %s, %s)
            ''', (
                f'Importado via Excel – {periodo}',
                json.dumps(config_data),
                json.dumps(candidates),
                0,
                0
            ))
            inserted += 1

        conn.commit()
        cur.close()
        conn.close()
        return inserted

    def get_history_report(self):
        history = self.get_voting_history()
        from collections import defaultdict, Counter

        all_candidates = []
        periods = []
        for entry in history:
            candidates = entry.get('candidates', [])
            if isinstance(candidates, str):
                import json as _json
                candidates = _json.loads(candidates)
            config = entry.get('config', {})
            if isinstance(config, str):
                import json as _json
                config = _json.loads(config)
            period = config.get('period', '') or entry.get('description', '')
            periods.append({
                'id': entry['id'],
                'description': entry['description'],
                'period': period,
                'timestamp': entry['timestamp'],
                'total_voters': entry.get('total_voters', 0),
                'total_votes': entry.get('total_votes', 0),
                'candidate_count': len(candidates),
            })
            for c in candidates:
                all_candidates.append({
                    'nome': c.get('nome', ''),
                    'categoria': c.get('categoria', ''),
                    'gestor': c.get('gestor', ''),
                    'vote_count': c.get('vote_count', 0),
                    'period': period,
                    'entry_id': entry['id'],
                })

        name_counter = Counter(c['nome'] for c in all_candidates)
        repeated = sorted(
            [{'nome': n, 'count': c, 'appearances': [
                {'period': x['period'], 'categoria': x['categoria'], 'votes': x['vote_count']}
                for x in all_candidates if x['nome'] == n
            ]} for n, c in name_counter.items() if c > 1],
            key=lambda x: -x['count']
        )

        cat_stats = defaultdict(int)
        for c in all_candidates:
            cat_stats[c['categoria']] += 1

        return {
            'periods': sorted(periods, key=lambda x: x['period']),
            'repeated': repeated,
            'cat_stats': dict(cat_stats),
            'total_entries': len(history),
            'total_candidates': len(all_candidates),
            'unique_candidates': len(set(c['nome'] for c in all_candidates)),
        }
