from flask import Flask, render_template, request, jsonify, redirect, url_for, send_from_directory, session, flash
import os
from datetime import datetime
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from PIL import Image
import io
import base64
from data_manager import DataManager
from functools import wraps

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SESSION_SECRET', 'dev-secret-key-change-in-production')
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

dm = DataManager()

# Credenciais de admin
ADMIN_USERNAME = 'leonardo'
ADMIN_PASSWORD = 'thaiane'

def require_admin_auth(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg', 'gif', 'xlsx'}

def process_image(image_file):
    img = Image.open(image_file)
    
    if img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info):
        img = img.convert('RGBA')
    else:
        img = img.convert('RGB')
    
    img.thumbnail((400, 400), Image.Resampling.LANCZOS)
    
    filename = secure_filename(f"{datetime.now().strftime('%Y%m%d%H%M%S')}.png")
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    if img.mode == 'RGBA':
        img.save(filepath, 'PNG')
    else:
        img.save(filepath, 'PNG')
    
    return filename

@app.route('/')
def index():
    return redirect(url_for('vote'))

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            return redirect(url_for('admin'))
        else:
            flash('Credenciais inválidas', 'error')
    
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    flash('Logout realizado com sucesso', 'success')
    return redirect(url_for('admin_login'))

@app.route('/admin')
@require_admin_auth
def admin():
    candidates = dm.get_candidates()
    config = dm.get_config()
    return render_template('admin.html', candidates=candidates, config=config)

@app.route('/admin/upload-excel', methods=['POST'])
@require_admin_auth
def upload_excel():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'Nenhum arquivo enviado'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Nenhum arquivo selecionado'})
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': 'Formato de arquivo inválido'})
        
        wb = load_workbook(filename=file.stream, read_only=True)
        ws = wb.active
        
        candidates = []
        if ws:
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:
                    continue
                
                colaborador = row[0]
                justificativa = row[1] if len(row) > 1 else ''
                gestor = row[2] if len(row) > 2 else ''
                periodo = row[3] if len(row) > 3 else 'Q3/2025'
                categoria = row[4] if len(row) > 4 else 'Eu Faço a Diferença'
                
                candidate = {
                    'nome': str(colaborador).strip(),
                    'justificativa': str(justificativa).strip() if justificativa else '',
                    'gestor': str(gestor).strip() if gestor else '',
                    'periodo': str(periodo).strip() if periodo else 'Q3/2025',
                    'categoria': str(categoria).strip() if categoria else 'Eu Faço a Diferença',
                    'photo': 'default_avatar.png'
                }
                candidates.append(candidate)
        
        dm.save_candidates([])
        dm.reset_voting()
        
        for candidate in candidates:
            dm.add_candidate(candidate)
        
        return jsonify({'success': True, 'count': len(candidates)})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/upload-photo', methods=['POST'])
@require_admin_auth
def upload_photo():
    try:
        candidate_id = int(request.form.get('candidate_id'))
        
        if 'photo' in request.files and request.files['photo'].filename:
            file = request.files['photo']
            filename = process_image(file)
            dm.update_candidate(candidate_id, {'photo': filename})
            return jsonify({'success': True, 'filename': filename})
        
        elif 'photo_base64' in request.form:
            photo_data = request.form.get('photo_base64', '')
            if photo_data and photo_data.startswith('data:image'):
                photo_data = photo_data.split(',')[1]
            
            if photo_data:
                image_data = base64.b64decode(photo_data)
                image_file = io.BytesIO(image_data)
                filename = process_image(image_file)
                dm.update_candidate(candidate_id, {'photo': filename})
                return jsonify({'success': True, 'filename': filename})
        
        return jsonify({'success': False, 'error': 'Nenhuma foto enviada'})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/update-candidate', methods=['POST'])
@require_admin_auth
def update_candidate():
    try:
        data = request.get_json()
        candidate_id = int(data.get('id'))
        
        update_data = {
            'nome': data.get('nome'),
            'justificativa': data.get('justificativa'),
            'gestor': data.get('gestor'),
            'categoria': data.get('categoria')
        }
        
        dm.update_candidate(candidate_id, update_data)
        return jsonify({'success': True})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/delete-candidate', methods=['POST'])
@require_admin_auth
def delete_candidate():
    try:
        data = request.get_json()
        candidate_id = int(data.get('id'))
        dm.delete_candidate(candidate_id)
        return jsonify({'success': True})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/set-voting-config', methods=['POST'])
@require_admin_auth
def set_voting_config():
    try:
        data = request.get_json()
        end_date = data.get('end_date')
        
        config = dm.get_config()
        config['voting_end_date'] = end_date
        config['voting_active'] = True
        dm.save_config(config)
        
        return jsonify({'success': True})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/stop-voting', methods=['POST'])
@require_admin_auth
def stop_voting():
    try:
        config = dm.get_config()
        config['voting_active'] = False
        dm.save_config(config)
        return jsonify({'success': True})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/reset-voting', methods=['POST'])
@require_admin_auth
def reset_voting():
    try:
        dm.reset_voting()
        return jsonify({'success': True})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/reset-results', methods=['POST'])
@require_admin_auth
def reset_results():
    try:
        dm.reset_voting()
        return jsonify({'success': True})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/delete-all-candidates', methods=['POST'])
@require_admin_auth
def delete_all_candidates():
    try:
        dm.delete_all_candidates()
        return jsonify({'success': True})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/candidate/<int:candidate_id>')
@require_admin_auth
def get_candidate(candidate_id):
    try:
        candidates = dm.get_candidates()
        candidate = next((c for c in candidates if c['id'] == candidate_id), None)
        if candidate:
            return jsonify(candidate)
        else:
            return jsonify({'error': 'Candidato não encontrado'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/vote')
def vote():
    if not dm.is_voting_active():
        return render_template('voting_closed.html')
    
    candidates = dm.get_candidates()
    config = dm.get_config()
    
    professional = [c for c in candidates if 'Líder' not in c.get('categoria', '')]
    leader = [c for c in candidates if 'Líder' in c.get('categoria', '')]
    
    # Ordenar alfabeticamente por nome
    professional.sort(key=lambda x: x['nome'])
    leader.sort(key=lambda x: x['nome'])
    
    return render_template('vote.html', 
                         professional=professional, 
                         leader=leader, 
                         config=config)

@app.route('/submit-vote', methods=['POST'])
def submit_vote():
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        candidate_id = data.get('candidate_id')
        
        if not email.endswith('@claro.com.br'):
            return jsonify({'success': False, 'error': 'E-mail deve ser do domínio @claro.com.br'})
        
        if not dm.is_voting_active():
            return jsonify({'success': False, 'error': 'Votação encerrada'})
        
        candidates = dm.get_candidates()
        candidate = next((c for c in candidates if c['id'] == candidate_id), None)
        
        if not candidate:
            return jsonify({'success': False, 'error': 'Candidato não encontrado'})
        
        categoria = candidate.get('categoria', '')
        is_leader = 'Líder' in categoria
        
        # Verificar se já votou nesta categoria
        previous_vote_id = dm.has_voted(email, candidate_id)
        
        # Adicionar voto ao candidato (add_voter já remove voto anterior se existir)
        dm.add_vote(candidate_id)
        
        if not dm.add_voter(email, candidate_id):
            return jsonify({'success': False, 'error': 'Erro ao registrar voto'})
        
        return jsonify({'success': True})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/results')
def results():
    results_data = dm.get_results()
    config = dm.get_config()
    
    return render_template('results.html', 
                         results=results_data, 
                         config=config,
                         is_active=dm.is_voting_active())

@app.route('/api/results')
def api_results():
    results_data = dm.get_results()
    
    all_candidates = results_data['all_candidates']
    
    candidates_data = []
    for candidate in all_candidates:
        candidates_data.append({
            'id': candidate['id'],
            'nome': candidate['nome'],
            'vote_count': candidate['vote_count'],
            'categoria': candidate['categoria'],
            'photo': candidate['photo']
        })
    
    return jsonify({
        'candidates': candidates_data,
        'is_active': dm.is_voting_active()
    })

@app.route('/api/config')
def api_config():
    config = dm.get_config()
    return jsonify(config)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
